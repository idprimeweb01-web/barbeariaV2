"""
Utilitário de relatórios customizáveis.
- COLUNAS: catálogo com flag 'obrigatorio' — coluna obrigatória não pode ser excluída.
- gerar_dados(): busca agendamentos com joins e extrai apenas as colunas selecionadas.
- gerar_excel() / gerar_pdf(): retornam BytesIO prontos para stream HTTP.
"""
import io
from datetime import date
from collections import OrderedDict

from app.labels import L
from app.utils.tz import naive_brasilia


# ── Catálogo de colunas ────────────────────────────────────────────────────────

COLUNAS: OrderedDict = OrderedDict([
    ('data',        {'label': 'Data/Hora',          'obrigatorio': True}),
    ('cliente',     {'label': L('cliente'),         'obrigatorio': True}),
    ('servico',     {'label': L('servicos'),        'obrigatorio': True}),
    ('status',      {'label': 'Status',             'obrigatorio': True}),
    ('valor_total', {'label': f'{L("receita")} (R$)', 'obrigatorio': True}),
    ('barbeiro',    {'label': L('profissional'),    'obrigatorio': False}),
    ('metodo',      {'label': 'Método Pagto.',      'obrigatorio': False}),
    ('duracao',     {'label': 'Duração (min)',      'obrigatorio': False}),
    ('is_plano',    {'label': f'Via {L("plano")}', 'obrigatorio': False}),
])

COLUNAS_OBRIGATORIAS = {k for k, v in COLUNAS.items() if v['obrigatorio']}
COLUNAS_VALIDAS      = set(COLUNAS.keys())


def validar_colunas(colunas_solicitadas: list[str]) -> list[str]:
    """
    Valida e ordena colunas conforme o catálogo.
    Levanta ValueError se coluna inválida ou se coluna obrigatória for omitida.
    """
    invalidas = set(colunas_solicitadas) - COLUNAS_VALIDAS
    if invalidas:
        raise ValueError(f'Colunas inválidas: {", ".join(sorted(invalidas))}')

    omitidas_obrigatorias = COLUNAS_OBRIGATORIAS - set(colunas_solicitadas)
    if omitidas_obrigatorias:
        raise ValueError(
            f'Colunas obrigatórias não podem ser removidas: {", ".join(sorted(omitidas_obrigatorias))}'
        )

    # Mantém ordem canônica do catálogo
    return [k for k in COLUNAS if k in colunas_solicitadas]


# ── Busca de dados ─────────────────────────────────────────────────────────────

def gerar_dados(barbearia_id: int, de: date, ate: date, colunas: list[str]) -> list[dict]:
    """
    Retorna lista de dicts com apenas as colunas solicitadas.
    Sempre filtra por barbearia_id (isolamento tenant).
    """
    from sqlalchemy.orm import selectinload
    from app.extensions import db
    from app.models import (
        Agendamento, AgendamentoServico, Servico,
        Cliente, Barbeiro, Usuario,
    )

    ags = (
        Agendamento.query
        .options(selectinload(Agendamento.itens).selectinload(AgendamentoServico.servico))
        .filter(
            Agendamento.barbearia_id == barbearia_id,
            db.func.date(Agendamento.data_hora) >= de,
            db.func.date(Agendamento.data_hora) <= ate,
        )
        .order_by(Agendamento.data_hora)
        .all()
    )

    if not ags:
        return []

    # Batch: clientes e barbeiros (+ usuário do barbeiro) usados no relatório inteiro.
    clientes = {}
    if 'cliente' in colunas:
        clientes = {c.id: c for c in Cliente.query.filter(
            Cliente.id.in_({ag.cliente_id for ag in ags})).all()}

    barbeiro_nomes = {}
    if 'barbeiro' in colunas:
        barbeiros = {b.id: b for b in Barbeiro.query.filter(
            Barbeiro.id.in_({ag.barbeiro_id for ag in ags})).all()}
        usuarios = {u.id: u for u in Usuario.query.filter(
            Usuario.id.in_({b.usuario_id for b in barbeiros.values()})).all()} if barbeiros else {}
        barbeiro_nomes = {
            bid: (usuarios.get(b.usuario_id).nome if usuarios.get(b.usuario_id) else '—')
            for bid, b in barbeiros.items()
        }

    resultado = []
    for ag in ags:
        linha: dict = {}

        if 'data' in colunas:
            linha['data'] = ag.data_hora.strftime('%d/%m/%Y %H:%M')

        if 'cliente' in colunas:
            cli = clientes.get(ag.cliente_id)
            linha['cliente'] = cli.nome if cli else '—'

        if 'servico' in colunas:
            nomes = [it.servico.nome + (' [P]' if it.is_plano else '') for it in ag.itens if it.servico]
            linha['servico'] = ', '.join(nomes) if nomes else '—'

        if 'status' in colunas:
            linha['status'] = ag.status

        if 'valor_total' in colunas:
            linha['valor_total'] = float(ag.valor_total)

        if 'barbeiro' in colunas:
            linha['barbeiro'] = barbeiro_nomes.get(ag.barbeiro_id, '—')

        if 'metodo' in colunas:
            linha['metodo'] = ag.metodo_pagamento or '—'

        if 'duracao' in colunas:
            linha['duracao'] = ag.duracao_minutos

        if 'is_plano' in colunas:
            linha['is_plano'] = any(it.is_plano for it in ag.itens)

        resultado.append(linha)

    return resultado


# ── Exportação Excel ───────────────────────────────────────────────────────────

def gerar_excel(
    dados: list[dict],
    colunas: list[str],
    barbearia_nome: str,
    periodo: str,
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Relatório'

    # ── Cabeçalho do documento ────────────────────────────────────────────────
    ws.merge_cells(f'A1:{chr(64 + len(colunas))}1')
    ws['A1'] = barbearia_nome
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{chr(64 + len(colunas))}2')
    ws['A2'] = f'Relatório de {L("agendamento")}s — {periodo}'
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A3:{chr(64 + len(colunas))}3')
    ws['A3'] = f'Gerado em: {naive_brasilia().strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(size=9, color='888888')
    ws['A3'].alignment = Alignment(horizontal='right')

    # ── Linha de rótulos ──────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='1A1A2E')
    thin_border = Border(
        bottom=Side(style='thin'),
        right=Side(style='thin'),
    )
    for col_idx, col_key in enumerate(colunas, start=1):
        cell = ws.cell(row=5, column=col_idx, value=COLUNAS[col_key]['label'])
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # ── Linhas de dados ───────────────────────────────────────────────────────
    for row_idx, linha in enumerate(dados, start=6):
        for col_idx, col_key in enumerate(colunas, start=1):
            val = linha.get(col_key, '')
            if isinstance(val, bool):
                val = 'Sim' if val else 'Não'
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F2F2F2')
            cell.border = thin_border

    # ── Linha de totais ───────────────────────────────────────────────────────
    if dados and 'valor_total' in colunas:
        total_row = len(dados) + 6
        vt_col = colunas.index('valor_total') + 1
        total = sum(float(l.get('valor_total', 0)) for l in dados)
        ws.cell(row=total_row, column=vt_col - 1 if vt_col > 1 else vt_col,
                value='TOTAL:').font = Font(bold=True)
        tc = ws.cell(row=total_row, column=vt_col, value=total)
        tc.font = Font(bold=True)
        tc.number_format = 'R$ #,##0.00'

    # ── Larguras ──────────────────────────────────────────────────────────────
    larguras = {
        'data': 18, 'cliente': 25, 'servico': 35, 'status': 18,
        'valor_total': 16, 'barbeiro': 22, 'metodo': 15, 'duracao': 14, 'is_plano': 12,
    }
    for col_idx, col_key in enumerate(colunas, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = larguras.get(col_key, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Exportação PDF ─────────────────────────────────────────────────────────────

def gerar_pdf(
    dados: list[dict],
    colunas: list[str],
    barbearia_nome: str,
    periodo: str,
) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    orientacao = landscape(A4) if len(colunas) > 5 else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=orientacao,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    elementos = []

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    titulo_style = ParagraphStyle(
        'Titulo',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1A1A2E'),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        'Sub',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#444444'),
        spaceAfter=2,
    )
    data_style = ParagraphStyle(
        'DataGer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        spaceAfter=12,
    )

    elementos.append(Paragraph(barbearia_nome, titulo_style))
    elementos.append(Paragraph(f'Relatório de {L("agendamento")}s — {periodo}', sub_style))
    elementos.append(Paragraph(
        f'Gerado em: {naive_brasilia().strftime("%d/%m/%Y %H:%M")}',
        data_style,
    ))

    # ── Tabela ────────────────────────────────────────────────────────────────
    header = [COLUNAS[k]['label'] for k in colunas]
    linhas_tabela = [header]

    for linha in dados:
        row = []
        for col_key in colunas:
            val = linha.get(col_key, '')
            if isinstance(val, bool):
                val = 'Sim' if val else 'Não'
            elif isinstance(val, float):
                val = f'R$ {val:.2f}'
            row.append(str(val))
        linhas_tabela.append(row)

    # Linha de total
    if dados and 'valor_total' in colunas:
        total = sum(float(l.get('valor_total', 0)) for l in dados)
        total_row = [''] * len(colunas)
        total_row[colunas.index('valor_total')] = f'TOTAL: R$ {total:.2f}'
        if 'status' in colunas:
            total_row[colunas.index('status')] = f'{len(dados)} registros'
        linhas_tabela.append(total_row)

    # Larguras proporcionais
    larguras_pdf = {
        'data': 3.5, 'cliente': 4.5, 'servico': 6.0, 'status': 3.0,
        'valor_total': 3.0, 'barbeiro': 4.0, 'metodo': 2.5, 'duracao': 2.2, 'is_plano': 2.0,
    }
    col_widths = [larguras_pdf.get(k, 3.0) * cm for k in colunas]

    tabela = Table(linhas_tabela, colWidths=col_widths, repeatRows=1)
    tabela.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A2E')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        # Dados
        ('FONTNAME',   (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE',   (0, 1), (-1, -2), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        # Linha de total
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8E8E8')),
        # Grade
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabela)
    elementos.append(Spacer(1, 0.5 * cm))
    elementos.append(Paragraph(
        f'Total de registros: {len(dados)}',
        ParagraphStyle('rodape', parent=styles['Normal'], fontSize=8, textColor=colors.grey),
    ))

    doc.build(elementos)
    buf.seek(0)
    return buf
